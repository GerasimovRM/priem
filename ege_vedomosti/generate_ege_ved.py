import itertools
import json
from copy import copy
import re

import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.cell import Cell, MergedCell
from openpyxl.descriptors.excel import CellRange
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension




with open("generate_data.json", encoding="utf-8") as json_data:
    generate_data = json.load(json_data)

zz = generate_data["page_row_count"]

abiturients_data = {}
data_xlsx = load_workbook("data.xlsx")
sheet = data_xlsx.worksheets[0]
next_row = None
current_row = 1
subjects = False
for i, row in enumerate(sheet.rows, current_row):
    if subjects:
        subject_data = []
        for j, cell in enumerate(row, 1):
            if isinstance(cell, Cell) and cell.value is not None:
                if cell.value == "Справка для личного дела абитуриента сформирована из ФИС ГИА и приема для образовательной организации:":
                    abiturients_data[current_abiturient]["subjects"].sort()
                    abiturients_data[current_abiturient]["subjects"] = list(k for k, _ in itertools.groupby(abiturients_data[current_abiturient]["subjects"]))
                    subjects = False
                    break
                subject_data.append(cell.value)
        if subjects and subject_data:
            if "subjects" not in abiturients_data[current_abiturient]:
                abiturients_data[current_abiturient]["subjects"] = []
            abiturients_data[current_abiturient]["subjects"].append(subject_data)
    else:
        for j, cell in enumerate(row, 1):
            if isinstance(cell, Cell) and cell.value is not None:
                # print(cell, cell.value)
                if cell.value == "о результатах единого государственного экзамена":
                    current_abiturient = sheet.cell(i + 2, j).value
                    if current_abiturient not in abiturients_data:
                        abiturients_data[current_abiturient] = {}
                elif "документ, удостоверяющий личность:" in cell.value:
                    if "passport" not in abiturients_data[current_abiturient]:
                        abiturients_data[current_abiturient]["passport"] = cell.value.split(":")[-1].strip()
                elif cell.value == "Наименование предмета":
                    subjects = True
                    break
print(*sorted(abiturients_data.items(), key=lambda x: x[0]), sep='\n')

template = load_workbook("template.xlsx")
ws_template = template.worksheets[0]

result_xlsx = Workbook()
result_xlsx.create_sheet("Лист1")
ws_res = result_xlsx.worksheets[0]

for idx, cd in ws_template.column_dimensions.items():
     ws_res.column_dimensions[idx] = copy(cd)

no_fill = openpyxl.styles.PatternFill(fill_type=None)
side = openpyxl.styles.Side(border_style=None)
no_border = openpyxl.styles.borders.Border(
    left=side,
    right=side,
    top=side,
    bottom=side,
)

# Loop through all cells in all worksheets
for row in ws_res:
    for cell in row:
        # Apply colorless and borderless styles
        cell.fill = no_fill
        cell.border = no_border


for abit_i, current_abiturient in enumerate(list(abiturients_data.keys())):
    for i, row in enumerate(ws_template.rows, 1):
        for j, col in enumerate(row, 1):
            template_value = ws_template.cell(i, j).value
            if template_value is not None:
                if "<template: fio>" in template_value:
                    template_value = template_value.replace("<template: fio>", current_abiturient)
                elif "<template: pasport_data>" in template_value:
                    template_value = template_value.replace("<template: pasport_data>", abiturients_data[current_abiturient]["passport"])
                elif "<template: date>" in template_value:
                    template_value = template_value.replace("<template: date>",
                                                            generate_data["date"])
                elif "<template: employee1_fio>" in template_value:
                    template_value = template_value.replace("<template: employee1_fio>",
                                                            generate_data["employee1_fio"])
                elif "<template: employee1_position>" in template_value:
                    template_value = template_value.replace("<template: employee1_position>",
                                                            generate_data["employee1_position"])
                elif "<template: employee2_fio>" in template_value:
                    template_value = template_value.replace("<template: employee2_fio>",
                                                            generate_data["employee2_fio"])
                elif "<template: employee2_position>" in template_value:
                    template_value = template_value.replace("<template: employee2_position>",
                                                            generate_data["employee2_position"])

            ws_res.cell(i + zz * abit_i, j).value = template_value
            if ws_template.cell(i, j).has_style:
                ws_res.cell(i + zz * abit_i, j).font = copy(ws_template.cell(i, j).font)
                ws_res.cell(i + zz * abit_i, j).border = copy(ws_template.cell(i, j).border)
                ws_res.cell(i + zz * abit_i, j).fill = copy(ws_template.cell(i, j).fill)
                ws_res.cell(i + zz * abit_i, j).number_format = copy(ws_template.cell(i, j).number_format)
                ws_res.cell(i + zz * abit_i, j).protection = copy(ws_template.cell(i, j).protection)
                ws_res.cell(i + zz * abit_i, j).alignment = copy(ws_template.cell(i, j).alignment)

    for elem in ws_template.merged_cells:
        rows = list(elem.rows)[0]
        row = rows[0][0] + zz * abit_i
        cols = list(map(lambda x: x[1], rows))
        ws_res.merge_cells(start_row=row, end_row=row, start_column=cols[0], end_column=cols[-1])
        #print(row, cols)
        #print("=" * 20)

    for i, row in enumerate(ws_template.rows, 1):
        for j, col in enumerate(row, 1):
            template_value = ws_template.cell(i, j).value
            if template_value is not None:
                if "<template: exams_data>" in template_value:
                    head_i = i - 1
                    for ii in range(len(abiturients_data[current_abiturient]["subjects"])):
                        ws_res.merge_cells(start_row=i + ii + zz * abit_i, end_row=i + ii + zz * abit_i, start_column=1, end_column=2)
                        ws_res.merge_cells(start_row=i + ii + zz * abit_i, end_row=i + ii + zz * abit_i, start_column=5, end_column=6)
                        ws_res.merge_cells(start_row=i + ii + zz * abit_i, end_row=i + ii + zz * abit_i, start_column=7, end_column=8)
                        subject = abiturients_data[current_abiturient]["subjects"][ii]
                        ws_res.cell(i + ii + zz * abit_i, 1, subject[0])
                        ws_res.cell(i + ii + zz * abit_i, 3, subject[1])
                        ws_res.cell(i + ii + zz * abit_i, 4, subject[2])
                        ws_res.cell(i + ii + zz * abit_i, 5, subject[3])
                        if len(subject) > 4:
                            ws_res.cell(i + ii + zz * abit_i, 7, subject[4])
                        for k in range(1, 11):
                            if ws_template.cell(head_i, k).has_style:
                                ws_res.cell(i + ii + zz * abit_i, k).font = copy(ws_template.cell(head_i, k).font)
                                ws_res.cell(i + ii + zz * abit_i, k).border = copy(ws_template.cell(head_i, k).border)
                                ws_res.cell(i + ii + zz * abit_i, k).fill = copy(ws_template.cell(head_i, k).fill)
                                ws_res.cell(i + ii + zz * abit_i, k).number_format = copy(ws_template.cell(head_i, k).number_format)
                                ws_res.cell(i + ii + zz * abit_i, k).protection = copy(ws_template.cell(head_i, k).protection)
                                ws_res.cell(i + ii + zz * abit_i, k).alignment = copy(ws_template.cell(head_i, k).alignment)
                    break
    for idx, rd in ws_template.row_dimensions.items():
        ws_res.row_dimensions[idx + zz * abit_i] = copy(rd)

result_xlsx.save("result.xlsx")
