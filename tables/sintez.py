from openpyxl import load_workbook

# Load in the workbook
from openpyxl.cell import MergedCell

wb1 = load_workbook('./Физика ФПММ.xlsx')

# Get sheet names
sheet1 = wb1["Лист1"]
i = 1
students = []
while True:
    if sheet1.cell(i, 1).value is None:
        break
    # print(sheet1.cell(i, 1).value)
    students.append(sheet1.cell(i, 1).value.lower().replace("\xa0", " "))
    i += 1

with open("bad.txt", encoding="utf-8") as bad_in:
    bad = list(map(lambda x: x.strip().lower(), bad_in.readlines()))
print(bad)

wb2 = load_workbook("./23.07.2022.xlsx")
sheet2 = wb2["Лист1"]
result = []
with open("result.txt", "w", encoding="utf-8") as output:
    for i in range(1, 1600):
        cell_value = sheet2.cell(i, 2)
        if isinstance(cell_value, MergedCell):
            continue
        cell_value = cell_value.value
        if cell_value is None:
            continue
        cell_value = cell_value.lower()
        if cell_value in students and cell_value not in bad:
            students.remove(cell_value)
            subres = [sheet2.cell(i, j).value for j in range(1, 19)]
            print(*map(lambda x: "-" if x is None else x, subres), sep='\t', file=output)

    # result.append()
    # print(cell_value)

